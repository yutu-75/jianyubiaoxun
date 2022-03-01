import re
import yaml
import time
import json
import base64
import random
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from hyper.contrib import HTTP20Adapter


from yubiao.images_code import img_main
from yubiao.chaojiying import main_cjy

# from Common.dir_config import *
url = 'https://www.jianyu360.com/jylab/supsearch/index.html'


# 打开配置文件数据
with open("config.yaml", 'r', encoding="UTF-8") as fs:
    data_yaml = fs.read()
    # print(data_yaml,type(data_yaml))
    # print(yaml.load(data_yaml, Loader=yaml.FullLoader)['config'])
    data = yaml.load(data_yaml, Loader=yaml.FullLoader)['config']



# cookie池
with open('cookie.json','r',encoding='utf-8') as f:
    cookie_list = json.loads(f.read())

# cookie = random.choice(cookie_list)

# 由于cookie池的cookie太少，只有一个cookie能用就没有使用cookie池
cookie = 'SESSIONID=5c89fdbcce799db8a7bf48317553ebe53753ad0c; SESSIONID=5c89fdbcce799db8a7bf48317553ebe53753ad0c; UM_distinctid=17f253b90a0e-0a20d9790d84d5-a3e3164-1fa400-17f253b90a1de6; selectNum=1; ud_safe=Ql1BWwFGUwdXB0ILFllSRwYEUgZGWUQN; limitSearchTextFlag=jnLid1645597524978894335; Hm_lvt_52c42de35032567eb9d7a24a43c84bda=1645597135,1645672913,1645682084; c__utmc=875156445.624666141; CNZZDATA1261815924=974490152-1645591097-https%253A%252F%252Fopen.jianyu360.com%252F%7C1645685430; c__utma=875156445.624666141.281617177.1645689006.1645694485.6; Hm_lpvt_52c42de35032567eb9d7a24a43c84bda=1645695899; c__utmb=875156445.624666141.1645694485.1645695899.3; userid_secure=naBVKCbyLjTl1Bv9CFTonTMXo9S/iBO96N3VUoUsa+wW3ABlZ0dgf2GehQxhmhHkJuSNvccp0Nptlidi5INj+sVU+25pmQYLUvKB/5H9WoVuVN8B++mf2tFaA9fPH3r04jH5mNk7I2BKbQLQ2Fba2VnVio0dthQGfLgAIbEUS2ug3De0+YAzY+jYRqazos9mhvYUBlslelcVjStxbEqMzYP+mWCB8BcCfVCJHVm4Vda5tmwN99NBOqyzCQxe40MO+C2glCCy1W8hDAztlmYPnbqgxClI5cqS2a0nGlQL4kxrCYnUH1POwgt8ncNKPET2fC5ve5EHO5q9b1cksccnNCoqKjIwMjItMDItMjAgMDA6MDA6MDA='


class jianyu:
    api_url = 'https://www.jianyu360.cn/front/pcAjaxReq'
    cookie_list = []
    headers = {
        "accept": "*/*",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8",
        "content-length": "273",
        "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        "cookie": cookie,
        "sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
        "sec-ch-ua-mobile": '?0',
        "sec-ch-ua-platform": "Windows",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",

        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
        "x-requested-with": "XMLHttpRequest",
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
    }
    cookie_false = []       # 失效cookie
    cache_list = []  # 缓存数据以便判断是否达到最大页 从而终止
    data = None
    def __init__(self):
        with open("config.yaml", 'r', encoding="UTF-8") as f1:
            data_yaml = f1.read()
            # print(data_yaml,type(data_yaml))
            # print(yaml.load(data_yaml, Loader=yaml.FullLoader)['config'])
            self.data = yaml.load(data_yaml, Loader=yaml.FullLoader)['config']
        with open('cookie.json', 'r', encoding='utf-8') as f:
            self.cookie_list = json.loads(f.read())


    # cookie随机
    def cookie_choice(self):
        cookie_r = random.choice(list(set(self.cookie_list)-set(self.cookie_false)))
        self.headers['cookie'] = cookie_r

    # excel 创建excel表
    def excel_start(self):

        wb = Workbook()
        # 激活 worksheet
        ws = wb.active
        wb.save('./yubiao_excel/{}.xlsx'.format(self.data['searchvalue']))
        wb.close()

    # excel表处理
    def excel_wirte(self, list: list):
        wb = load_workbook('./yubiao_excel/{}.xlsx'.format(self.data['searchvalue']))

        # wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 60
        ws['A1'] = '序号'
        ws['B1'] = '省份'
        ws['C1'] = '城市'
        ws['D1'] = '公告标题'
        ws['E1'] = '公告类别'
        ws['F1'] = '发布时间'
        ws['G1'] = '采购单位'
        ws['H1'] = '中标单位'
        ws['I1'] = '中标金额(元)'
        ws['J1'] = '项目名称'
        ws['K1'] = '剑鱼标讯地址'
        ws.append(list)
        wb.save('./yubiao_excel/{}.xlsx'.format(self.data['searchvalue']))
        wb.close()

    # 验证码处理
    def code_testing(self, page_text):
        # print(page_text)
        print('遇到验证码啦！！！         ε(┬┬﹏┬┬)3')
        print('正在处理验证码中          (oﾟvﾟ)ノ')

        # base64_str1 = re.findall(r'base64,(.*?)"', page_text)  # 获取图片加密后的字符串
        # if not base64_str1:
        #     base64_str1 = page_text['imgData']
        # code_name = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>', page_text)  # 获取需要识别的图片文字

        self.img_merge(page_text['imgData'], page_text['textVerify'])
        pic_str = main_cjy()  # 调用打码平台
        data = {
            "antiVerifyCheck": pic_str,
            "imgw": "331"
        }
        headers_1 = {
            #
            # ":authority": "www.jianyu360.com",
            # ":method": "POST",
            # ":path": "/article/content/ABCY1wIdSkoMysvEn9zcE8zJicCCj1mdmB0Py8wOCE3eFVzcSdUCbA%3D.html",
            # ":scheme": "https",
            "accept": "application/json,text/javascript,*/*;q=0.01",
            "accept-encoding": "gzip,deflate,br",
            "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,ar;q=0.7",
            "app": "jyweb",
            "content-length": "55",
            "content-type": "application/x-www-form-urlencoded;charset=UTF-8",
            "cookie": cookie,
            "origin": "https://www.jianyu360.com",
            # "referer": "https://www.jianyu360.com/article/content/ABCY1wIdSkoMysvEn9zcE8zJicCCj1mdmB0Py8wOCE3eFVzcSdUCbA%3D.html",
            "sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
            "x-requested-with": "XMLHttpRequest"
        }
        # sessions = requests.session()
        # sessions.mount("https://www.jianyu360.com", HTTP20Adapter())  # 解决headers报错问题
        qwq = requests.post(url=self.api_url, headers=self.headers, data=data).json()
        print('验证码已处理完啦！       (≧∇≦)ﾉ！！！')
        # print(qwq)
        return qwq       # 验证验证结果

    # 图片处理

    @staticmethod
    def img_merge( base64_str, code_name):
        with open('./yubiao/img/code.png', 'wb') as f_png:
            """
            base64_str: 加密后的图片字符串
            code_name: 需要识别的文字
            """
            # 图片base64+加字符更改解密
            # print(base64_str)
            print("需要识别处理的文字是:",code_name)
            img_data = base64.b64decode(base64_str)
            f_png.write(img_data)  # 保存原始解密后的图片
            img_main(code_name)  # 调用普通合成处理

    def main_start(self):

        # 自定义搜索关键字
        # name_text = input("请输入你需要的搜索的内容:")
        # name_text = '乡村振兴规划编制'
        # if name_text:
        #     self.data['searchvalue'] =  name_text


        pageNumber = 1
        num = 1
        self.excel_start()
        while True:
            print("第{}页的数据：".format(pageNumber))
            # print(self.data)
            # print(self.headers)
            # print(self.data)
            self.data['pageNumber'] = pageNumber
            page = requests.post(url=self.api_url, headers=self.headers, data=self.data)
            try:

                page = page.json()

                if not page['list']:
                    return
            except Exception as e:
                print("error:",e)
                for i_index in range(5):
                    page1 = self.code_testing(page)
                    if page1.get('antiVerify') == 1:
                        time.sleep(1)
                        page = requests.post(url=self.api_url, headers=self.headers, data=data).json()
                        break
                else:
                    print("验证码无法验证！")
                    time.sleep(3)
                    return

            cache_list_Sub = []  # 缓存数据以便判断是否达到最大页 从而终止
            # 数据信息获取
            for i in page['list']:
                area = i.get('area', 'null')  # 省份
                city = i.get('city', 'null')  # 城市
                title = i.get('title', 'null')  # 标题
                subtype = i.get('subtype', 'null')  # 公告类别
                publishtime = datetime.fromtimestamp(i['publishtime'])  # 发布时间
                buyer = i.get('buyer', 'null')  # 采购单位
                winner = i.get('winner', 'null')  # 中标单位
                budget = i.get('budget', 'null')  # 中标金额(万元)
                projectname = i.get('projectname', 'null')  # 项目名称
                url_jy = 'https://www.jianyu360.com/article/content/' + i.get('_id') + '.html'  # 剑鱼标讯地址
                total_list = [num, area, city, title, subtype, publishtime, buyer, winner, budget, projectname,
                              url_jy]
                if total_list[1:] in self.cache_list:
                    return
                else:
                    cache_list_Sub.append(total_list[1:] )
                self.excel_wirte(total_list)

                print("第{}条的数据：".format(num))
                print(total_list)
                num += 1
            self.cache_list = cache_list_Sub
            pageNumber += 1



a = jianyu()
a.main_start()
