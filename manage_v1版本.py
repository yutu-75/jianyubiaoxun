import requests
import re
import datetime
import time
import random
import base64
import openpyxl
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook
from chaojiying import main
from img2 import img_main
from yubiao.code import code_testing
url = 'https://www.jianyu360.com/jylab/supsearch/index.html'


# 获取本地cookie
with open("cookie.txt",'r',encoding='utf-8') as f_c:
    cookie = f_c.read()



# session = requests.Session()
# session.get(url="http://www.xiladaili.com")
# # 去代理商获取url
# response = session.get(
#     url='http://www.xiladaili.com/api/?uuid=9e2ed14b846a4d28ad3afb8442916f54&num=6&place=中国&protocol=2&port=80&sortby=0&repeat=1&format=3&position=1')
# a = response.text
# ip_list = a.split(' ')  # 代理ip列表
# print(ip_list)







name_text = input("请输入你需要的搜索的内容:")
# name_text = '乡村振兴规划编制'





def excel_start():
    from openpyxl import Workbook
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active
    wb.save('./yubiao_excel/{}.xlsx'.format(name_text))
    wb.close()




# excel表处理
def excel_wirte(list):
    wb = load_workbook('./yubiao_excel/{}.xlsx'.format(name_text))
    # wb = Workbook()
    ws = wb.active
    ws['A1'] = '序号'
    ws['B1'] = '省份'
    ws['C1'] = '城市'
    ws['D1'] = '公告标题'
    ws['E1'] = '公告类别'
    ws['F1'] = '发布时间'
    ws['G1'] = '采购单位'
    ws['H1'] = '中标单位'
    ws['I1'] = '中标金额(万元)'
    ws['J1'] = '项目名称'
    ws['K1'] = '剑鱼标讯地址'
    ws.append(list)
    wb.save('./yubiao_excel/{}.xlsx'.format(name_text))
    wb.close()


def qwq(q):
    if len(q) == 0:
        return ''
    else:
        return q[0]

num_c = 0
# 数据处理
def data_handle(url_str_list):
    global num_c
    for i in url_str_list:

        num_c += 1
        txt = []
        txt.append(num_c)                                         # 序号
        if 'area' not in i:
            i['area'] = ' '
        txt.append(i['area'])                        # 省份
        if 'city' not in i:
            i['city'] = ' '
        txt.append(i['city'])                        # 城市
        if 'title' not in i:
            i['title'] = ' '
        txt.append(i['title'])                       # 公告标题
        if 'subtype' not in i:
            i['subtype'] = ' '
        txt.append(i['subtype'])                     # 公告类别

        dateArray = datetime.datetime.fromtimestamp(int(i['publishtime']))
        otherStyleTime = dateArray.strftime("%Y-%m-%d %H:%M:%S")
        txt.append(otherStyleTime)                              # 发布时间

        # i['_id'] = 'ABCY1wIcj0%2FIDk7Ent4cAcOMzAZCSdgZnx1KS8FJi8Nd2BzfAVUCRQ%3D'
        url = 'https://www.jianyu360.com/article/content/{}.html'.format(i['_id'])
        page_text = requests.get(url=url, headers=headers,).text
        try:

            a = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>',page_text)
            # print(a)
            if len(a) > 0:
                raise
        except:
            print('错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了错了')
            # print(page_text)
            code_testing()
            page_text = requests.get(url=url, headers=headers).text

        buyer_ = re.findall(r'var buyer_ = "(.*?)";', page_text)
        # print(buyer_)
        txt.append(qwq(buyer_))                                 # 采购单位

        s_winner = re.findall(r'var s_winner = "(.*?)"', page_text)
        # print(s_winner)
        txt.append(qwq(s_winner))                              # 中标单位

        bidamount = re.findall(r'var bidamount_class=(.*?)\r\n', page_text)
        # print(bidamount)
        txt.append(qwq(bidamount))                              # 中标金额(万元)

        projectname_ = re.findall(r'var projectname_ = "(.*?)";', page_text)
        # print(projectname_,)
        txt.append(qwq(projectname_))                           # 项目名称

        txt.append(url)                                         # 剑鱼标讯地址
        print(txt)
        # print(url)
        print('数据保存中******')


        excel_wirte(txt)
        time.sleep(10)




headers = {
# ":authority": "www.jianyu360.com",
# ":method": "GET",
# ":path": "/jylab/supsearch/index.html",
# ":scheme": "https",
"accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",

"accept-encoding": "gzip, deflate, br",
"accept-language": "zh-CN,zh;q=0.9,en;q=0.8,ar;q=0.7",
"cache-control": "max-age=0",
"cookie": cookie,
"sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
"sec-ch-ua-mobile": '?0',
"sec-fetch-dest": "document",
"sec-fetch-mode": "navigate",
"sec-fetch-site": "none",
"sec-fetch-user": "?1",
"upgrade-insecure-requests": '1',
'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
    # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
}

data_start = {                        # 搜素的数据量小于等于两页 用此data           默认使用，首次搜索比用此data
    "keywords": name_text,
    "publishtime": '',
    "timeslot": '',
    "area": "",
    "subtype":"" ,
    "minprice": "",
    'maxprice': "",
    'industry': "",
    'selectType': 'title',






}


page = requests.post(url=url,headers=headers,data=data_start).text

# print(page+'*******************************************************')
data_str2 = re.findall(r'secondList = \[(.*?)\r\nif', page)                 # 从混淆代码中获取对应的数据  获取第二页数据
if data_str2:                                                               # 判断是否有分页
    excel_start()
    """
    此网站获取的数据量小于或等于一百会直接把数据发生完，为此不需要去获取分页信息
    """
    data_str = re.findall(r'var list = \[(.*?)]:"";', page)                 # 从混淆代码中获取对应的数据

    data_list = eval('['+data_str[0].replace(']!=null?[',',')+']')[0:50]    # 第一页数据

    data_str2 = re.findall(r'secondList = \[(.*?)\r\nif', page)             # 从混淆代码中获取对应的数据


    data_list2 = eval('['+data_str2[0].replace("]']",']'))[0:50]            # 第二页数据

    data_list = data_list+data_list2                                        # 合并数据

    # print(data_list)
    data_handle(data_list)

else:
    # print('qwq')

    while True:
        try:
            start_time = input("请输入开始时间,时间格式为 (2020-02-02):")
            start_time = start_time + ' 00:00:00'
            timeArray = time.strptime(start_time, "%Y-%m-%d %H:%M:%S")
            start_time = int(time.mktime(timeArray))

            end_time = input("请输入结束时间,时间格式为 (2020-02-20):")
            end_time = end_time + ' 23:59:59'
            # print(end_time)
            timeArray = time.strptime(end_time, "%Y-%m-%d %H:%M:%S")
            end_time = int(time.mktime(timeArray))

            publishtime = str(start_time) + "_" + str(end_time)
            break
        except:
            print('时间格式错误！')


    txt = []
    co = code_testing()


    # 判断是不是检测
    # if co == 'ok':
    page = requests.post(url=url, headers=headers, data=data_start).text
    #
    #
    # # print(page)
    data_str = re.findall(r'var list = \[(.*?)]:"";', page)  # 从混淆代码中获取对应的数据
    # print(data_str)
    #
    if "]!=null?[" not in data_str:                         # 判断是否有内容
    #
        excel_start()
    #     # print('qwq1111')
    #     data_list = eval('[' + data_str[0].replace(']!=null?[', ',') + ']')  # 第一页数据
    #     # print(data_list)
    #     num_len = len(data_list)/2
    #     data_list = data_list[0:int(num_len)]
    #     # print(data_list)
    #     # for i in data_list:
    #     #     print(i)
    #     # print(len(data_list))
    #     txt += data_list
        sign = True
        num = 1
        while sign:
            # print(qwq)

            data_add = {  # 搜素的数据量超过两页 用此data                 跳转页数使用
                "pageNumber": num,
                "reqType": "bidSearch",
                "searchvalue": name_text,
                "area": "",
                "subtype": "",
                "publishtime": publishtime,
                'selectType': 'title',
                "minprice": "",
                'maxprice': "",
                'industry': "",
                "tabularflag": "Y",
            }
            num += 1
            time.sleep(1)

            try:

                page = requests.post(url='https://www.jianyu360.com/front/pcAjaxReq', headers=headers,
                                     data=data_add)
                # print(page.text)
                a = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>', page.text)
                page = page.json()
                # break

                if a:
                    print(a)
                if len(a) > 0:
                    raise 1
            except:
                print('遇到验证码了******************')
                code_testing()
                ci = 0
                while True:

                    try:
                        page = requests.post(url='https://www.jianyu360.com/front/pcAjaxReq', headers=headers, data=data_add).json()

                        break
                    except:

                        code_testing()# 人居环境 2021-02-
                        if ci == 5:
                            print('账号被封禁！ 换账号吧！')
                            # break
                        ci += 1

                # print(page.text)
                # print(page.json())
                # page = page.json()
            data_list = page['list']
            if page['list'] in txt:
                break
            try:
                txt += page['list']
                print("获取分页链接数据:",len((txt)))
                if len(txt) >= 200 and len(txt) % 200 == 0:
                    judge = input("是否终止获取页面？不终止任意输入 输入 y 终止:")
                    if judge == 'y' or judge == 'Y':
                        print('qwq')

                        sign = False

                # data_handle(txt)

                time.sleep(1)
            except:
                print('获取完成！')
                break
        print(len(txt),"条数据")
        print("跳转到子页面获取数据.....")

        data_handle(txt)
            # print(txt)

            # break
    else:
        print("抱歉没有内容！")










# # 验证码图片处理
# def img(base64_str,code_name):
#     with open('../img/code.png', 'wb') as f:
#         # print(base64_str)
#         imgdata = base64.b64decode(base64_str[0].replace('&#43;', '+'))
#         f.write(imgdata)
#         img_main(code_name)
# # print(requests.post(url=url, headers=headers, data=data).text)
# a = requests.post(url=url, headers=headers, data=data).text
# print(a)
# url_z = 'https://www.jianyu360.com/article/content/{}.html'
#
#
# str1 = re.findall(r'"_id":"(.*?)","area',a)
# base64_str = re.findall(r'base64,(.*?)"', a)
# code_name = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>', a)

# img(base64_str, code_name)
# pic_str = main()
#
# data = {
#     "antiVerifyCheck": pic_str,
#     "imgw": "331",
# }
# headers_1 = {
#
#     "accept": "application / json, text / javascript, * / *; q = 0.01",
#     "accept-encoding": 'gzip, deflate, br',
#     'accept-language': 'zh - CN, zh;',
#
#     "content-length": "54",
#     "content-type": "application / x - www - form - urlencoded;charset = UTF - 8",
#     "cookie": 'selectNum=1; SESSIONID=c432393c01d63048b98572119bccae8e5f75638f; SESSIONID=c432393c01d63048b98572119bccae8e5f75638f; UM_distinctid=1783dc350d6739-0cdde569b60b06-5771031-1fa400-1783dc350d77c1; selectNum=1; Hm_lvt_72331746d85dcac3dac65202d103e5d9=1615943755,1615945011,1615998044,1616030209; CNZZDATA1261815924=2030409669-1615938790-%7C1616030606; Hm_lpvt_72331746d85dcac3dac65202d103e5d9=1616034094; userid_secure=Gl3YiszD0hfEHkdl94ufrPwiJSJL1QDz8iWHddpgx6tSgdFNQWXcDc3jSr6Y4zGDKNRsNtdg4SgF5kSaNasWzjd4Hf2W1L0JNNrHiFG7q0kYjrA5gwHymxLX7L38/2hAvwH82V6SK8HEaPmv3vZerQATetkQQ6844Pvw/hrrwN11Ihu1KjsgR2Smxw/ZXDqTVNuMEkkknNXwLiOY7gh+A7B+15MD5s5NhvCGaCvdEiQxMNvY8jbLvHqJbnUVHhibkNepcW7fO4rq+IH0odB6CilZmgFHxrPkTJOlTWp/sLTQMOmFBSjX00NE9KSd3v2qrTn8bXPqI2LhdVL8yN6mnioqKjIwMjEtMDMtMTQgMDA6MDA6MDA=',
#     "sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
#     "sec-ch-ua-mobile": '?0',
#     "sec-fetch-dest": "empty",
#     "sec-fetch-mode": "cors",
#     "sec-fetch-site": "same-origin",
#     'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
#
#     "with": "XMLHttpRequest",
#
# }
# qwq = requests.post(url=url, headers=headers_1, data=data).text
# print(qwq)
# 有序去重
# url_str_list = sorted(set(str1),key=str1.index)








# if not url_str_list:
#     img(base64_str, code_name)
#     pic_str = main()
#
#
# num = 0
# url_str_list = url_str_list[0:]
# for i in url_str_list:
#
#     txt = []
#     url = 'https://www.jianyu360.com/article/content/{}.html'.format(i)
#
#     print(url)
#     # params = {
#     #     "kds": "政府网站与政务媒体 + 政府 + 网站 + 政务 + 媒体",
#     #     "industry": "信息技术",
#     # }'
#     page_text = requests.get(url=url,headers=headers).text
#     base64_str = re.findall(r'base64,(.*?)"', page_text)
#     code_name = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>', page_text)
#     print(code_name,":code_name")
#     # print(page_text)
#     # print(base64_str)
#     if base64_str:
#         print('被检测了！')
#         print('解决检测中。。。')
#         img(base64_str,code_name)
#         pic_str = main()
#         print(page_text)
#
#         # code_page = requests.get(url=url, headers=headers).text
#         # base64_str = re.findall(r'base64,(.*?)"', page_text)
#         #
#         # code_name = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>',page_text)
#
#         data = {
#             "antiVerifyCheck": pic_str,
#             "imgw": "331",
#         }
#         headers_1 = {
#
#             "accept": "application / json, text / javascript, * / *; q = 0.01",
#             "accept-encoding": 'gzip, deflate, br',
#             'accept-language': 'zh - CN, zh;',
#
#             "content-length": "54",
#             "content-type": "application / x - www - form - urlencoded;charset = UTF - 8",
#             "cookie": 'selectNum=1; SESSIONID=c432393c01d63048b98572119bccae8e5f75638f; SESSIONID=c432393c01d63048b98572119bccae8e5f75638f; UM_distinctid=1783dc350d6739-0cdde569b60b06-5771031-1fa400-1783dc350d77c1; selectNum=1; Hm_lvt_72331746d85dcac3dac65202d103e5d9=1615943755,1615945011,1615998044,1616030209; CNZZDATA1261815924=2030409669-1615938790-%7C1616030606; Hm_lpvt_72331746d85dcac3dac65202d103e5d9=1616034094; userid_secure=Gl3YiszD0hfEHkdl94ufrPwiJSJL1QDz8iWHddpgx6tSgdFNQWXcDc3jSr6Y4zGDKNRsNtdg4SgF5kSaNasWzjd4Hf2W1L0JNNrHiFG7q0kYjrA5gwHymxLX7L38/2hAvwH82V6SK8HEaPmv3vZerQATetkQQ6844Pvw/hrrwN11Ihu1KjsgR2Smxw/ZXDqTVNuMEkkknNXwLiOY7gh+A7B+15MD5s5NhvCGaCvdEiQxMNvY8jbLvHqJbnUVHhibkNepcW7fO4rq+IH0odB6CilZmgFHxrPkTJOlTWp/sLTQMOmFBSjX00NE9KSd3v2qrTn8bXPqI2LhdVL8yN6mnioqKjIwMjEtMDMtMTQgMDA6MDA6MDA=',
#             "sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
#             "sec-ch-ua-mobile": '?0',
#             "sec-fetch-dest": "empty",
#             "sec-fetch-mode": "cors",
#             "sec-fetch-site": "same-origin",
#             'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
#
#             "with": "XMLHttpRequest",
#
#         }
#         requests.get(url=url, headers=headers, data=data)
#     time.sleep(3)
#
#     # print(page_text)
#     parser = etree.HTMLParser(encoding="utf-8")
#     tree = etree.HTML(page_text,parser)
#     num += 1
#     txt.append(num)
#     project_name = tree.xpath('//*[@id="tab2"]/div[2]/table/tr[2]/td[4]/div[1]/text()')[0]
#     txt.append(project_name)
#     tree.xpath('//*[@id="tab2"]/div[2]/table/tr[1]/td[4]/text()')
#     city = tree.xpath('//*[@id="tab2"]/div[2]/table/tr[1]/td[4]/text()')
#     if city:
#         txt.append(city[0])
#     else:
#         txt.append('')
#     Bid_winner = tree.xpath('//*[@id="tab2"]/div[4]/table/tr[1]/td[2]/div[1]/text()')
#     if Bid_winner:
#         txt.append(Bid_winner[0])
#     else:
#         txt.append('')
#     # Bid_amount = tree.xpath('//*[@id="bidInfoCont"]/table/tr[1]/td[4]/div[1]//text()')
#     Bid_amount = re.findall(r'var _bidamount = (.*?);',page_text)[0]
#     txt.append(Bid_amount)
#
#     # 时间戳格式转换
#     # print(page_text)
#     time_project = re.findall(r'mytime = (.*?) ;', page_text)
#     if time_project:
#         time_project = time_project[0]
#     else:
#         print('no time')
#     dateArray = datetime.datetime.fromtimestamp(int(time_project))
#     otherStyleTime = dateArray.strftime("%Y-%m-%d %H:%M:%S")
#     time1 = otherStyleTime
#     txt.append(time1)
#     txt.append(url)
#     # print(time1)
#     print(txt)
#     excel_wirte(txt)
#     time.sleep(1)
#     break
#
#
#
# print(response.text)
# print(response.text)
# print(response.status_code)
# import time
# time.sleep(1)
# print(response.elapsed)
# response.encoding = 'utf-8'
# response = response.text
# with open('qwq.html','w',encoding='utf-8') as f:
#     f.write(response)
#
# print(response)
# 解决不规范网站
# parser = etree.HTMLParser(encoding="utf-8")
# tree = etree.HTML(response, parser=parser)
# r_list = tree.xpath('/html/body/footer/div[1]/div/ul//text()')
# print(r_list)


