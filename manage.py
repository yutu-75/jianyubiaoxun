import requests
import re
from datetime import datetime
import time
import json
import base64
from openpyxl import load_workbook
import random
from hyper.contrib import HTTP20Adapter
from openpyxl import Workbook

from yubiao.images_code import img_main
from yubiao.chaojiying import main
url = 'https://www.jianyu360.com/jylab/supsearch/index.html'

# 打开cookie池
with open('cookie.json','r',encoding='utf-8') as f:
    cookie_list = json.loads(f.read())
# cookie = random.choice(cookie_list)
cookie = 'selectNum=1; UM_distinctid=1795502b69bbbe-063c120965fd2e-d7e1739-1fa400-1795502b69c4d9; SESSIONID=4400b5fe9305474c7883a592e2e16b50f5e3ca63; SESSIONID=4400b5fe9305474c7883a592e2e16b50f5e3ca63; selectNum=1; Hm_lvt_72331746d85dcac3dac65202d103e5d9=1623308801,1623373260; article=/article/content/ABCZnZdczI%2FGSs4Emd1c2I4CjA4QCBmYXh%2FKygoKk0kDw0IVg%3D%3D; CNZZDATA1261815924=625946359-1620628612-%7C1623378849; Hm_lpvt_72331746d85dcac3dac65202d103e5d9=1623379193; userid_secure=GluR70yLkksTGD+f+Up4BpTNXg8K9Cw2RPxjiMdAaBIGi9Lg/NPNfXiMC5uNnkSVKLLUVPWBZx3rQ6F0qMf62DXOm9EBpe4gJrfuE1qALTa+O04vmNkqRrFoD7SwNFUd4azlHgAE7T1eAfakoAV4/5AQeyuSKPL6EGn98PbWN9DMkRWJTX0vTddiW3m+ZFQqZQftiMF8dLwRH31nf1a79gabsIdPVaDPowbmp5OtFoWwukzVDZygRb23UWzkgWlQHoYMlj0HgyLVaK1VP7hXhZbW1YhsqXfuUfe79wzaFk/1cRtPz+83Uw+ORIwH+T7dcJJdR7DYW5OQGiWfAErXkioqKjIwMjEtMDYtMDYgMDA6MDA6MDA='
# print(cookie)

# session = requests.Session()
# session.get(url="http://www.xiladaili.com")
# # 去代理商获取url
# response = session.get(
#     url='http://www.xiladaili.com/api/?uuid=9e2ed14b846a4d28ad3afb8442916f54&num=6&place=中国&protocol=2&port=80&sortby=0&repeat=1&format=3&position=1')
# a = response.text
# ip_list = a.split(' ')  # 代理ip列表
# print(ip_list)




class jianyu:
    api_url = 'https://www.jianyu360.com/front/pcAjaxReq'
    publishtime = ''
    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,ar;q=0.7",
        "cache-control": "max-age=0",
        "cookie": cookie,
        "sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
        "sec-ch-ua-mobile": '?0',
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36',
    }
    name_text = ''

    def excel_start(self):

        wb = Workbook()
        # 激活 worksheet
        ws = wb.active
        wb.save('./yubiao_excel/{}.xlsx'.format(self.name_text))
        wb.close()

    # excel表处理
    def excel_wirte(self,list:list):
        wb = load_workbook('./yubiao_excel/{}.xlsx'.format(self.name_text))
        # wb = Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '省份'
        ws['C1'] = '城市'
        ws['D1'] = '公告标题'
        ws['E1'] = '公告类别'
        ws['F1'] = '发布时间'
        ws['G1'] = '采购单位'
        ws['H1'] = '中标单位'
        ws['I1'] = '中标金额(万元)'
        ws['J1'] = '项目名称'
        ws['K1'] = '剑鱼标讯地址'
        ws.append(list)
        wb.save('./yubiao_excel/{}.xlsx'.format(self.name_text))
        wb.close()

    # 验证码处理
    def code_testing(self, page_text):
        # print(page_text)
        print('遇到验证码啦！！！         ε(┬┬﹏┬┬)3')
        print('正在处理验证码中          (oﾟvﾟ)ノ')
        base64_str1 = re.findall(r'base64,(.*?)"', page_text)  # 获取图片加密后的字符串
        if not base64_str1:

            base64_str1 = page_text['imgData']
        code_name = re.findall(r'<div>请在下图依次点击：<span>(.*?)</span></div>', page_text)  # 获取需要识别的图片文字
        self.img(base64_str1,code_name)
        pic_str = main()  # 调用打码平台
        data = {
            "antiVerifyCheck": pic_str,
            "imgw": "331"
        }
        headers_1 = {
            #
            # ":authority": "www.jianyu360.com",
            # ":method": "POST",
            # ":path": "/article/content/ABCY1wIdSkoMysvEn9zcE8zJicCCj1mdmB0Py8wOCE3eFVzcSdUCbA%3D.html",
            # ":scheme": "https",
            "accept": "application/json,text/javascript,*/*;q=0.01",
            "accept-encoding": "gzip,deflate,br",
            "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,ar;q=0.7",
            "app": "jyweb",
            "content-length": "55",
            "content-type": "application/x-www-form-urlencoded;charset=UTF-8",
            "cookie": cookie,
            "origin": "https://www.jianyu360.com",
            # "referer": "https://www.jianyu360.com/article/content/ABCY1wIdSkoMysvEn9zcE8zJicCCj1mdmB0Py8wOCE3eFVzcSdUCbA%3D.html",
            "sec-ch-ua": '"Google Chrome";v="89", "Chromium";v="89", ";Not A Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
            "x-requested-with": "XMLHttpRequest"
        }
        # sessions = requests.session()
        # sessions.mount("https://www.jianyu360.com", HTTP20Adapter())  # 解决headers报错问题
        qwq = requests.post(url=self.api_url, headers=headers_1, data=data)
        print('验证码已处理完啦！       (≧∇≦)ﾉ！！！')
        print(qwq.text)
        return qwq.json()                         # 验证验证结果

    # 图片处理
    def img(self,base64_str, code_name):
        with open('./yubiao/img/code.png', 'wb') as f_png:
            """
            base64_str: 加密后的图片字符串
            code_name: 需要识别的文字
            """
            # 图片base64+加字符更改解密
            imgdata = base64.b64decode(base64_str[0].replace('&#43;', '+'))
            f_png.write(imgdata)  # 保存原始解密后的图片
            img_main(code_name)  # 调用普通合成处理

    def main(self):
        name_text = input("请输入你需要的搜索的内容:")
        self.name_text = name_text
        # name_text = '乡村振兴规划编制'
        print('不填回车则跳过！')

        while True:
            try:
                start_time = input("请输入开始时间,时间格式为 (2020-02-02):")
                if len(start_time) < 1:
                    break
                start_time = start_time + ' 00:00:00'
                timeArray = time.strptime(start_time, "%Y-%m-%d %H:%M:%S")
                start_time = int(time.mktime(timeArray))

                end_time = input("请输入结束时间,时间格式为 (2020-02-20):")
                end_time = end_time + ' 23:59:59'
                # print(end_time)
                timeArray = time.strptime(end_time, "%Y-%m-%d %H:%M:%S")
                end_time = int(time.mktime(timeArray))

                publishtime = str(start_time) + "_" + str(end_time)
                print(publishtime)
            except:
                print('输入错误！')

        pageNumber = 1
        num = 1
        self.excel_start()
        while True:
            data = {
                'pageNumber': pageNumber,
                'reqType': 'bidSearch',
                'searchvalue': name_text,
                'area': '',
                'subtype': '',
                'publishtime': '',
                'selectType': 'title',
                'minprice': '',
                'maxprice': '',
                'industry': '',
                'tabularflag': 'Y',
            }


            page = requests.post(url=self.api_url, headers=self.headers, data=data)
            try:

                page = page.json()
                if not len(page['list']):
                    return
            except:
                page.encoding = 'utf-8'
                page = page.text
                # print(page)
                for i in range(5):
                    page1 = self.code_testing(page)
                    if page1.get('antiVerify') == 1:
                        page = requests.post(url=self.api_url, headers=self.headers, data=data).json()
                        break
                print("验证码无法验证！")
                time.sleep(3)
                return
                # print(page)

            # 数据信息获取
            for i in page['list']:
                area = i.get('area', 'null')  # 省份
                city = i.get('city', 'null')  # 城市
                title = i.get('title', 'null')  # 标题
                subtype = i.get('subtype', 'null')  # 公告类别
                publishtime = datetime.fromtimestamp(i['publishtime'])  # 发布时间
                buyer = i.get('buyer', 'null')  # 采购单位
                winner = i.get('winner', 'null')  # 中标单位
                bidamount = i.get('bidamount', 'null')  # 中标金额(万元)
                projectname = i.get('projectname', 'null')  # 项目名称
                url_jy = 'https://www.jianyu360.com/article/content/' + i.get('_id') + '.html'  # 剑鱼标讯地址
                total_list = [num, area, city, title, subtype, publishtime, buyer, winner, bidamount, projectname,
                              url_jy]
                self.excel_wirte(total_list)
                print(total_list)
                num += 1
            pageNumber += 1
            # print(page)
            # break

a = jianyu()
a.main()






